import datetime as dt
from collections import OrderedDict

import openpyxl as px
from CoreApp.services.access import (KvantObjectExistsMixin,
                                     KvantTeacherAndAdminAccessMixin)
from CoreApp.services.utils import ObjectManipulationManager
from DiaryApp.models import KvantLesson, KvantTaskBase
from django.urls import reverse_lazy as rl
from LoginApp.models import KvantUser
from openpyxl.utils import get_column_letter
from openpyxl.writer.excel import save_virtual_workbook
from RegisterApp.models import StaffPersonalInfo, StudentPersonalInfo
from RegisterApp.serializers import (StaffPersonalInfoSerializer,
                                     StudentPersonalInfoSerializer)

from .models import KvantCourse, KvantCourseType


class KvantAdminAccessMixin(KvantTeacherAndAdminAccessMixin):
    def _permissionTest(self, user):
        return user.permission == 'Администратор'


class KvantUserDeleteAccessMixin(KvantAdminAccessMixin, KvantObjectExistsMixin):
    request_object_arg = 'user_identifier'

    def _objectExiststTest(self, object_id):
        return KvantUser.objects.filter(id=object_id).exists()


class KvantCourseTypeAccessMixin(KvantObjectExistsMixin):
    request_object_arg = 'subject_identifier'

    def _objectExiststTest(self, object_id):
        return KvantCourseType.objects.filter(id=object_id).exists()


class KvantCourseAccessMixin(KvantObjectExistsMixin):
    request_object_arg = 'course_identifier'

    def _objectExiststTest(self, object_id):
        return KvantCourse.objects.filter(id=object_id).exists()


class CourseSubjectManipulationManager(ObjectManipulationManager):
    def _constructRedirectUrl(self, **kwargs):
        return rl('subject_detail', kwargs={'subject_identifier': kwargs.get('obj').id})


class CourseManipulationManager(ObjectManipulationManager):    
    def updateObject(self, request):
        schedules = list(self.object.schedule.all())
        course_or_errors = self._getUpdatedObject(request)

        if isinstance(course_or_errors, KvantCourse):
            [schedule.delete() for schedule in schedules]
        return self.getResponse(course_or_errors)
    
    def _constructRedirectUrl(self, **kwargs):
        return rl('course_detail', kwargs={'course_identifier': kwargs.get('obj').id})


class PersonalInfoExcelImport:
    def __init__(self):
        self._wb = px.Workbook()
        self._ws = self._wb.active
    
    def importPersonalInfo(self, user_type, ids):
        if user_type == 'Ученик':
            return self._importUsersPersonalInfo(
                serializer=StudentPersonalInfoSerializer,
                users=StudentPersonalInfo.objects.filter(user__permission=user_type, user__id__in=ids))
        return self._importUsersPersonalInfo(
            serializer=StaffPersonalInfoSerializer,
            users=StaffPersonalInfo.objects.filter(user__permission=user_type, user__id__in=ids)
        )
    
    def _importUsersPersonalInfo(self, users, serializer):
        if not users.exists():
            return save_virtual_workbook(self._wb) 
        
        self._writeTable(self._parseTableHead(serializer(users.first()).data), 1)
        
        for i in range(len(users)):
            self._writeTable(self._parseTableData(serializer(users[i]).data), i + 2)
        return save_virtual_workbook(self._wb)

    def _writeTable(self, data, row):
        for col in range(len(data)):
            self._ws.cell(row=row, column=col + 1, value=data[col])
            self._ws.column_dimensions[get_column_letter(col + 1)].width = 30

    def _parseTableHead(self, data):
        table_head = []
        for col_name in data.keys():
            if isinstance(data[col_name], OrderedDict):
                generated_cols = self._parseTableHead(data[col_name])
                table_head += [f"{col_name} {sub_col_name}" for sub_col_name in generated_cols]
            else:
                table_head += [col_name]
        return table_head
    
    def _parseTableData(self, data):
        table_head = []
        for col_name in data.keys():
            if isinstance(data[col_name], OrderedDict):
                table_head += self._parseTableData(data[col_name])
            else:
                table_head += [data[col_name]]
        return table_head


class GenerateRegisterLink(ObjectManipulationManager):
    def __init__(self, forms, object=None):
        self._out_string = str()
        super(GenerateRegisterLink, self).__init__(forms, object)
    
    def createRegisterLink(self, request):
        links_or_errors = self._getCreatedObject(request)

        if not isinstance(links_or_errors, list):
            return '\n'.join([error.as_text() for error in links_or_errors.values()])
        
        for link_obj in links_or_errors:
            self._out_string += request.build_absolute_uri(f"{rl('register_page')}?key={link_obj.key}") + '\n'
        return self._out_string

    def _constructRedirectUrl(self, **kwargs): return


class GenerateCourseLessons:
    def __init__(self, course):
        self._course = course
    
    def generateLessons(self, start_date, end_date):
        start_date = dt.datetime.strptime(start_date, '%Y-%m-%d') 
        end_date = dt.datetime.strptime(end_date, '%Y-%m-%d')
        
        for schedule in self._course.schedule.all():
            days_delta = self._getDaysDelta(start_date, self._getWeekDayNumber(schedule.week_day))
            
            lesson = start_date + dt.timedelta(days=days_delta); i = 0
            if lesson > end_date or lesson < start_date:
                return False
            
            while start_date <= lesson <= end_date:
                self._createLesson(lesson, f'Урок на {lesson.date()} #{i}', self._course, schedule.time)
                lesson = lesson + dt.timedelta(days=7); i += 1
        return True

    def _getWeekDayNumber(self, weekday):
        return {'ПН': 0, 'ВТ': 1,
                'СР': 2, 'ЧТ': 3,
                'ПТ': 4, 'СБ': 5,
                'ВС': 6}.get(weekday)
    
    def _getDaysDelta(self, start_day, lesson_day):
        if start_day.weekday() >= lesson_day:
            return (lesson_day + 7) - start_day.weekday()
        return lesson_day - start_day.weekday()
    
    def _createLesson(self, date, name, course, time):
        return KvantLesson.objects.create(
            date=date, base=self._createLessonBase(name), course=course, time=time)
    
    def _createLessonBase(self, name):
        return KvantTaskBase.objects.create(title=name)


def getCourseById(course_id):
    """ Возвращает курс по его course_id """
    return KvantCourse.objects.get(id=course_id)


def getCourseTypeById(subject_id):
    return KvantCourseType.objects.get(id=subject_id)


def getCourseQuery(user):
    """ Получает множество курсов основываясь на user """
    return {
        'Ученик': lambda user: KvantCourse.objects.filter(students__in=[user]),
        'Учитель': lambda user: KvantCourse.objects.filter(teacher=user),
        'Администратор': lambda user: KvantCourse.objects.all(),
    }[user.permission](user)


def getCourseTypeQuery(user):
    types = []
    for course in getCourseQuery(user):
        types.append(course.type.name)
    return set(types)


def getSubjectById(id):
    return KvantCourseType.objects.get(id=id)


def getSubjectData(subject):
    courses = KvantCourse.objects.filter(type=subject)
    return {
        'courses': courses,
        'courses_count': len(courses),
        'teachers': set([course.teacher for course in courses]),
        'teachers_count': len(set([course.teacher for course in courses])),
        'students_count': len(set([student for course in courses for student in course.students.all()])),
    }


def getCourseData(course):
    return {
        'subjects': allSubjects().exclude(id=course.type.id),
        'students': allUsers('Ученик').exclude(
            id__in=[student.id for student in course.students.all()]),
        'teachers': allUsers('Учитель').exclude(id=course.teacher.id),
    }


def deleteCourseLessons(course):
    [lesson.delete() for lesson in KvantLesson.objects.filter(course=course)]


def getSubjectGroups(subject):
    if subject == 'all':
        return set([course.name for course in allCourses()])
    return set([course.name for course in allCourses().filter(type=getSubjectById(subject))])


allCourses = lambda: KvantCourse.objects.all()
allSubjects = lambda: KvantCourseType.objects.all()
allUsers = lambda permission: KvantUser.objects.filter(permission=permission)
